# create MySQL DB
import pymysql.cursors
import openpyxl

# load WB with data
wb = openpyxl.load_workbook(file_name, data_only=True)
# load active sheet
sheet = wb['Sheet']

# create DB
connection = pymysql.connect(host=host,
                             user=user,
                             port=3306,
                             password=password,
                             db=db_name,
                             autocommit=True)
try:
    with connection.cursor() as cursor:
        # if table doesn't exist, create a new table
        create_table = "CREATE TABLE IF NOT EXISTS Produce (name varchar(30), " \
                       "cost_per_pound float, pounds_sold float, total float)"
        cursor.execute(create_table)

        # parse excel with required info
        for rowNum in range(2, sheet.max_row):  # skip the first row
            produceName = sheet.cell(row=rowNum, column=1).value
            produceCPP = sheet.cell(row=rowNum, column=2).value
            poundsSold = sheet.cell(row=rowNum, column=3).value
            total = sheet.cell(row=rowNum, column=4).value
            # collect all metrics in a list
            to_db = [produceName, produceCPP, poundsSold, total]
            # sql query to add row
            insert_new_row = "INSERT INTO Produce (name, cost_per_pound, pounds_sold, total) VALUES (%s, %s, %s, %s)"
            # add the new row
            cursor.execute(insert_new_row, to_db)

        # sql query to select all possible rows in the table
        sql_query = "SELECT * FROM Produce"

        # execute the query
        cursor.execute(sql_query)

        rows = cursor.fetchall()
        # print all table rows
        for row in rows:
            print(row)

finally:
    connection.close()
